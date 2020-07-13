import smtplib
import xlwings as xw
import time
import numpy as np
import os
import xlrd
import threading
import pythoncom
import win32com.client 
from xlutils.copy import copy
from shutil import copyfile
from datetime import datetime
from openpyxl import load_workbook
from email.header import Header
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication

def get_information(config_file):
    workbook = load_workbook(config_file)
    booksheet = workbook.worksheets[0]
    row_number = booksheet.max_row
    column_number = booksheet.max_column
    line = [[] for i in range(row_number-1)]
    
    for row in range(row_number-1):
        for col in range(column_number):
            line[row].append(booksheet.cell(row = row+2,column = col+1).value)
    return line

def update_excel(file_path,new_file_path):
    pythoncom.CoInitialize()
    lock.acquire()
    xlapp = win32com.client.DispatchEx("Excel.Application")
    wb = xlapp.Workbooks.Open(file_path)
    wb.RefreshAll() 
    time.sleep(5)
    wb.SaveAs(new_file_path) 
    xlapp.Quit() 
    lock.release()
    pythoncom.CoInitialize()
    
def copy_excel(config_file):

    app=xw.App(visible=False,add_book=False)
    app.screen_updating=False
    wb=app.books.open(config_file)
    marco1=wb.macro('Delcon')
    marco2=wb.macro('saveAsXlsx')
    marco1()
    #time.sleep(5)
    marco2()
    #time.sleep(5)
    wb.close()
    app.quit()

def daily_log(old_path,new_path):
    copyfile(old_path,new_path)
    os.remove(old_path)
    
def send_email(path,new_path,file_name,fromadd,toadd,ccadd,Sub,Body,Smtp,port):
    pythoncom.CoInitialize()
    lock.acquire()
    excelFile = file_name
    
    fromaddr = fromadd
    #password = pwd
    toaddrs = toadd.split(';')
    ccaddrs = ccadd.split(';')

    excelApart = MIMEApplication(open(path+'\\'+excelFile, 'rb').read())
    excelApart.add_header('Content-Disposition', 'attachment', filename=excelFile)

    m = MIMEMultipart()
    m.attach(MIMEText(Body, 'plain', 'utf-8'))
    m.attach(excelApart)
    m['Subject'] = Sub
    m['From'] = Header(fromaddr)
    m['To'] = Header(",".join(toaddrs))
    m['Cc'] = Header(",".join(ccaddrs))
 
    try:
        server = smtplib.SMTP(Smtp,port)
        #server.login(fromaddr,password)
        server.sendmail(fromaddr, toaddrs, m.as_string())
        print('success')
        server.quit()
        if os.path.exists(new_path+'\\'+excelFile) == False:
            daily_log(path+'\\'+excelFile,new_path+'\\'+excelFile)
        else:
            os.remove(new_path+'\\'+excelFile)
            daily_log(path+'\\'+excelFile,new_path+'\\'+excelFile)
    except smtplib.SMTPException as e:
        print('error:',e)
    lock.release()
    pythoncom.CoInitialize()
    

if __name__ == '__main__':
    datetime = datetime.now().date().isoformat()
    path = os.path.dirname(os.getcwd())
    data = get_information(path + '\\CONFIG\\config.xlsm')
    x = np.array(data)

    back_path = path+'\\BACKUP\\'+datetime
    if os.path.exists(back_path) == False:
        os.mkdir(back_path)
        
    threadt = []
    threadu = []
    lock = threading.Lock()
    
    for i in range(x.shape[0]):
        u = threading.Thread(target=update_excel,args=(path+'\\REPORT\\'+data[i][0],path+'\\COPY'+data[i][0]))
        t = threading.Thread(target=send_email,args=(path,back_path,data[i][8],data[i][3],data[i][4],data[i][5],data[i][1],data[i][2],data[i][6],data[i][7]))
        threadu.append(u)
        threadt.append(t)
    for thru in threadu:
        thru.start()
    thru.join()  
        
    copy_excel(path+'\\CONFIG\\config.xlsm')
    
    for i in range(x.shape[0]):
        os.remove(path+'\\COPY'+data[i][0])
    
    for thrt in threadt:
        thrt.start()
    
    thrt.join()